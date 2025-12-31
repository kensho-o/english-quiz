import random
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import List

import streamlit as st
from openpyxl import load_workbook

XLSX_PATH = Path("quiz.xlsx")
QUESTIONS_PER_RUN = 10


# =====================
# データ定義
# =====================
@dataclass
class QuizItem:
    id: str
    id_num: int
    ja: str
    cloze_en: str
    answer: str
    full_ja: str


def normalize(s: str) -> str:
    return str(s).strip().lower()


def build_full_en(cloze_en: str, answer: str) -> str:
    return cloze_en.replace("＿", "_").replace("____", answer)


# =====================
# Excel 読み込み
# =====================
def load_items_from_xlsx(path: Path) -> List[QuizItem]:
    if not path.exists():
        raise FileNotFoundError("quiz.xlsx が見つかりません（app.pyと同じ階層に置いてください）")

    wb = load_workbook(path, data_only=True)
    ws = wb.worksheets[0]

    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(v).strip() for v in header]
    idx = {name: i for i, name in enumerate(headers)}

    required = ["id", "ja", "cloze_en", "answer", "full_ja"]
    missing = [c for c in required if c not in idx]
    if missing:
        raise ValueError(f"Excelに必要な列がありません: {missing} / 現在: {headers}")

    def get(row, col):
        i = idx[col]
        v = row[i] if i < len(row) else None
        return "" if v is None else str(v).strip()

    items: List[QuizItem] = []
    bad = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        _id = get(row, "id")
        ja = get(row, "ja")
        cloze_en = get(row, "cloze_en").replace("＿", "_")
        answer = get(row, "answer")
        full_ja = get(row, "full_ja")

        # id を数値化
        try:
            id_num = int(_id)
        except:
            bad.append((_id, "id が数字ではありません"))
            continue

        if "____" not in cloze_en:
            bad.append((_id, "cloze_en に ____ がありません"))
            continue
        if not answer:
            bad.append((_id, "answer が空です"))
            continue
        if not full_ja:
            bad.append((_id, "full_ja が空です"))
            continue

        items.append(
            QuizItem(
                id=_id,
                id_num=id_num,
                ja=ja,
                cloze_en=cloze_en,
                answer=answer,
                full_ja=full_ja,
            )
        )

    st.session_state["bad_rows"] = bad
    return items


# =====================
# クイズ初期化（ID範囲指定）
# =====================
def init_quiz(min_id: int, max_id: int):
    items = load_items_from_xlsx(XLSX_PATH)

    pool = [it for it in items if min_id <= it.id_num <= max_id]

    if len(pool) < QUESTIONS_PER_RUN:
        raise ValueError(
            f"指定範囲（ID {min_id}〜{max_id}）の有効問題が {len(pool)} 件です。"
            f"{QUESTIONS_PER_RUN} 件以上必要です。"
        )

    quiz = random.sample(pool, QUESTIONS_PER_RUN)

    st.session_state.quiz = [asdict(q) for q in quiz]
    st.session_state.i = 0
    st.session_state.correct = 0
    st.session_state.wrong = 0
    st.session_state.skipped = 0
    st.session_state.phase = "question"
    st.session_state.last = None
    # NOTE: user_input はウィジェットが管理するので、ここでは触らない（エラー回避）


# =====================
# UI
# =====================
st.set_page_config(page_title="英単語クイズ", page_icon="📝")
st.title("📝 英単語クイズ")

if "phase" not in st.session_state:
    st.session_state.phase = "start"

# ---- サイドバー ----
with st.sidebar:
    st.header("出題範囲指定（ID）")

    min_id = st.number_input("下限ID", min_value=1, value=int(st.session_state.get("min_id", 1)), step=1, key="min_id")
    max_id = st.number_input("上限ID", min_value=1, value=int(st.session_state.get("max_id", 1000)), step=1, key="max_id")

    st.divider()

    if st.button("リセット"):
        st.session_state.clear()
        st.rerun()

    bad = st.session_state.get("bad_rows", [])
    if bad:
        with st.expander("読み込み時にスキップされた行"):
            for _id, reason in bad[:200]:
                st.write(f"- ID={_id}: {reason}")


# ---- 開始画面 ----
if st.session_state.phase == "start":
    st.write(f"ID **{min_id}〜{max_id}** の範囲から **10問ランダム出題**します。")

    if st.button("▶️ スタート", type="primary"):
        try:
            init_quiz(int(min_id), int(max_id))
            st.rerun()
        except Exception as e:
            st.error(str(e))


# ---- 問題表示 ----
elif st.session_state.phase == "question":
    quiz = st.session_state.quiz
    i = st.session_state.i
    q = quiz[i]

    st.subheader(f"Q{i+1}/10")
    st.write(f"**日本語**：{q['ja']}")
    st.write(f"**英文**：{q['cloze_en']}")

    # keyを付けて入力を保持（ただし同じ実行サイクルで上書きしない）
    user = st.text_input("空欄に入る語句（大小文字は無視）", key="user_input")

    col1, col2 = st.columns(2)

    if col1.button("送信", type="primary"):
        if user.strip() == "":
            st.session_state.skipped += 1
            correct = False
            is_skip = True
        else:
            is_skip = False
            correct = normalize(user) == normalize(q["answer"])
            if correct:
                st.session_state.correct += 1
            else:
                st.session_state.wrong += 1

        st.session_state.last = {
            "is_skip": is_skip,
            "correct": correct,
            "user": user,
            "answer": q["answer"],
            "full_en": build_full_en(q["cloze_en"], q["answer"]),
            "full_ja": q["full_ja"],
        }
        st.session_state.phase = "feedback"
        st.rerun()

    if col2.button("スキップ"):
        st.session_state.skipped += 1
        st.session_state.last = {
            "is_skip": True,
            "correct": False,
            "user": "",
            "answer": q["answer"],
            "full_en": build_full_en(q["cloze_en"], q["answer"]),
            "full_ja": q["full_ja"],
        }
        st.session_state.phase = "feedback"
        st.rerun()


# ---- 解答表示 ----
elif st.session_state.phase == "feedback":
    last = st.session_state.last
    i = st.session_state.i
    total = QUESTIONS_PER_RUN

    if last["is_skip"]:
        st.info("スキップ")
    elif last["correct"]:
        st.success("正解")
    else:
        st.error("不正解")
        if last["user"]:
            st.write(f"あなたの解答：{last['user']}")
        st.write(f"正解：{last['answer']}")

    st.divider()
    st.write("**英文（全文）**")
    st.write(last["full_en"])
    st.write("**日本語訳**")
    st.write(last["full_ja"])

    if st.button("次へ ▶️", type="primary"):
        # ウィジェットの値は「次へ」でクリア（同一実行サイクルで上書きしないため）
        st.session_state.user_input = ""

        st.session_state.i += 1
        if st.session_state.i >= total:
            st.session_state.phase = "done"
        else:
            st.session_state.phase = "question"
        st.rerun()


# ---- 結果 ----
elif st.session_state.phase == "done":
    st.subheader("結果")
    st.write(f"正解：{st.session_state.correct}")
    st.write(f"不正解：{st.session_state.wrong}")
    st.write(f"スキップ：{st.session_state.skipped}")

    if st.button("もう一回（別の10問）", type="primary"):
        # 次回開始時に入力を空にしておく
        st.session_state.user_input = ""
        st.session_state.phase = "start"
        st.rerun()
